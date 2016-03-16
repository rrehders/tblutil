#! /usr/bin/env python3
"""Module: Command line utility to perform manipulation of table type data files"""

import argparse
import warnings
import openpyxl
import csv
import re
import os

__author__ = 'rrehders'


class InvalidFileType(Exception):
    pass


class InvalidExcelColumn(Exception):
    pass


def getfiletype(fname):
    """
    Returns a standardized file type based on the extension of the filename provided
    :param fname: Sting containing the path and file
    :return: String of standardized file types {'excel', 'csv'} or False if an invalid file type
    """
    if re.search('\.xlsx$', fname, re.IGNORECASE):
        return 'excel'
    elif re.search('\.csv$', fname, re.IGNORECASE):
        return 'csv'
    else:
        raise InvalidFileType(fname)


def cvtcolsstrtoset(strcols):
    """
    Convert a string of numeric or letter columns to a set of numeric columns
    :param strcols: Original String to convert
    :return: set of integers corresponding to spreadsheet columns
    """
    cols = set()
    params = strcols.split(',')
    for param in params:
        if param.isdecimal():
            cols.add(int(param))
        elif param.isalpha():
            cols.add(openpyxl.utils.column_index_from_string(param))
        else:
            raise InvalidExcelColumn(param)
    return cols


def extractxltable(xlsheet, cols=set()):
    if len(cols) == 0:
        # Build lists of values for each row
        print('Extracting all columns')
        table = []
        for rowOfCellObjs in xlsheet:
            row = []
            for cellObj in rowOfCellObjs:
                row += [cellObj.value]
                print('.', end='')
            table += [row]
            print('')
    else:
        # Discard columns which are invalid
        cols = cols.intersection(range(xlsheet.max_column+1))
        print('Extracting columns: ' + str(cols))
        table = []
        for rowOfCellObjs in xlsheet:
            row = []
            col = 1
            for cellObj in rowOfCellObjs:
                if col in cols:
                    row += [cellObj.value]
                col += 1
                print('.', end='')
            table += [row]
            print('')
    return table


def extractlisttable(csvsheet, cols=set()):
    if len(cols) == 0:
        # Build lists of values for each row
        print('Extracting all columns')
        table = []
        for rowOfCells in csvsheet:
            row = []
            for cellObj in rowOfCells:
                row += [cellObj]
                print('.', end='')
            table += [row]
            print('')
    else:
        # Discard columns which are invalid
        cols = cols.intersection(range(len(csvsheet)+1))
        print('Extracting columns: ' + str(cols))
        table = []
        for rowOfCells in csvsheet:
            row = []
            col = 1
            for cellObj in rowOfCells:
                if col in cols:
                    row += [cellObj]
                col += 1
                print('.', end='')
            table += [row]
            print('')
    return table

def outcsvtable(table, fnamebase='/output.csv', subname=''):
    """
    :param table: Data to be written
    :param fnamebase: Path and base Filename
    :param subname: Name to be appended to the base
    :return:
    """
    # Set the output filename based on sheet name
    ofname = os.path.splitext(os.path.basename(fnamebase))[0] + '-' + subname + '.csv'

    # Open output file
    ofile = open(ofname, mode='w', newline='')

    # attach csv_writer to the output file
    ow = csv.writer(ofile)

    # Write out each row of the csv file
    print('Writing file ' + ofname)
    for i in range(len(table)):
        ow.writerow(table[i])
    print('')

    # Close output file
    ofile.close()


def tocsv(fname, sheetnum=-1, cols=set()):
    """
    Perform the conversion of a sheet from an excel file to a csv file
    :param fname: String containing the filename of the input Excel document
    :param sheetnum: Index of the sheet to be converted
    :param cols: Optional set listing the columns to be included
    :return: True if successful, None if incomplete
    """
    # validate file type
    try:
        if getfiletype(fname) != 'excel':
            raise InvalidFileType(fname)
    except InvalidFileType:
        print('ERR: '+fname+' is not a valid filetype to convert to csv')
        print('Valid filetypes are: .xls, .xlsx')
        return

    # load the target workbook
    print('Convert an Excel worksheet to CSV')
    try:
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(fname, data_only=True)
        sheetnms = wb.get_sheet_names()
    except Exception as err:
        print('ERR: '+fname+' '+str(err))
        return

    # Validate the sheetnum, -1 represents all sheets
    if sheetnum != -1:
        if sheetnum not in range(len(sheetnms)):
            # seek user input, provided sheetnum was invalid
            sheetnum = -1
            # Display Sheets in the workbook and ask which sheet to convert
            print('Sheets in '+fname)
            for i in range(len(sheetnms)):
                print(' | '+str(i)+' - '+sheetnms[i], end='')
            print(' |')

            # Get sheet selection
            while sheetnum not in range(len(sheetnms)):
                sheetnum = int(input('Convert which sheet ? '))
            print('')

    # Loop through the workbook's sheets
    for sheetidx in range(len(sheetnms)):
        # test each sheet to see if it should be converted and output
        if sheetnum < 0 or sheetidx == sheetnum:
            # Set the active sheet to the selection
            print('Sheet: '+sheetnms[sheetnum])
            xlsheet = wb.get_sheet_by_name(sheetnms[sheetnum])

            # extract cells from the input excel file and set of columns
            table = extractxltable(xlsheet, cols)

            # output a csv file of table with the hybrid file-sheet name
            outcsvtable(table, fname, sheetnms[sheetidx])

    # Completed all activities, function returns True to mark success
    return True

def extractcols(fname, cols, sheetnum=-1):
    """
    Create a new file of the same type as the input file but containing only the columns identified
    :param fname: String containing the input file path, name and extension
    :param cols: String containg the columns requested for extraction
    :return: True if successful, None if incomplete
    """
    # Validate colums are requested
    if not len(cols):
        print('ERR: no columns specified for extraction')
        return

    if getfiletype(fname) is 'excel':
        # load the target workbook
        print('Extract columns from an Excel worksheet')
        try:
            wb = openpyxl.load_workbook(fname, data_only=True)
            sheetnms = wb.get_sheet_names()
        except UserWarning:
            pass
        except Exception as err:
            print('ERR: '+fname+' '+str(err))
            return

        # Seek user input if sheetnum is not file contains more than one sheet
        if sheetnum not in range(len(sheetnms)):
            sheetnum = -1
            # Display Sheets in the workbook and ask which sheet to convert
            print('Sheets in '+fname)
            for i in range(len(sheetnms)):
                print(' | '+str(i)+' - '+sheetnms[i], end='')
            print(' |')

            # Get sheet selection
            while sheetnum not in range(len(sheetnms)):
                sheetnum = int(input('Convert which sheet ? '))
            print('')

        # Set the active sheet to the selection
        print('Sheet: '+sheetnms[sheetnum])
        xlsheet = wb.get_sheet_by_name(sheetnms[sheetnum])

        # extract cells from the input excel file and set of columns
        table = extractxltable(xlsheet, cols)

        # Set the output filename based on sheet name
        ofname = sheetnms[sheetnum]+'.csv'

        # Open output file
        ofile = open(ofname, mode='w', newline='')

        # attach csv_writer to the output file
        ow = csv.writer(ofile)

        # Write out each row of the csv file
        print('Writing file ' + ofname + '.')
        for i in range(len(table)):
            print('.', end='')
            ow.writerow(table[i])
        print('')

        # Close output file
        ofile.close()

    elif getfiletype(fname) is 'csv':
        print('Extract columns from an CSV File')
        with open(fname, 'r') as filein:
            csvin = csv.reader(filein)
            data = [row for row in csvin]

        table = extractlisttable(data, cols)
        # Set the output filename based on sheet name
        ofname = os.path.basename(fname)+'_extract.csv'

        # Open output file
        ofile = open(ofname, mode='w', newline='')

        # attach csv_writer to the output file
        ow = csv.writer(ofile)

        # Write out each row of the csv file
        print('Writing file ' + ofname + '.')
        for i in range(len(table)):
            print('.', end='')
            ow.writerow(table[i])
        print('')

        # Close output file
        ofile.close()

    # Completed all activities, function returns True to mark success
    return True


def cvtstrindextoset(strindex):
    if strindex is None or strindex == '':
        return {0,0}
    else:
        strindex = strindex.rstrip()

    if len(strindex.split(',')) == 1:
        return set(cvtcolsstrtoset(strindex), cvtcolsstrtoset(strindex))
    else:
        indexes = strindex.rstrip().split(',')[:2]
        return cvtcolsstrtoset(indexes)


def joinfiles(file1, file2, index):
    pass

def create_parser():
    parser = argparse.ArgumentParser(
        description='Perform various table manipulations'
    )

    parser.add_argument(
        "-c",
        "--cols",
        type=str,
        help='Optional value to specify which columns to use'
    )

    parser.add_argument(
        "-s",
        "--sheet",
        type=int,
        help='Optional value to specify which columns to use'
    )

    parser.add_argument(
        "-t",
        "--to",
        type=str,
        help='Optional value to specify the second file to use with join'
    )

    parser.add_argument(
        "-i",
        "--index",
        type=str,
        help='Optional value to specify the columns to use as index between two files for join'
    )

    parser.add_argument(
        'action',
        type=str,
        choices=['tocsv', 'extract', 'join'],
        help='Action to perform on the table'
    )

    parser.add_argument(
        'file',
        type=str,
        help='File containing the table'
    )

    return parser


def main():
    parser = create_parser()
    args = parser.parse_args()
    if args.action is 'tocsv':
        tocsv(args.file, args.sheet)
    elif args.action is 'extract':
        extractcols(args.file, args.cols)
    elif args.action is 'join:':
        joinfiles(args.file, args.to, cvtstrindextoset(args.index))
    else:
        pass

    print('Complete')

if __name__ == '__main__':
    main()

from tblutil import InvalidFileType, InvalidExcelColumn
from tblutil import getfiletype, cvtcolsstrtoset, extractxltable, extractlisttable, tocsv, extractcols
import unittest
import openpyxl
import os


class SubFunctionsTestCase(unittest.TestCase):
    def test_tblutil_getFileType_XLSX(self):
        self.assertTrue(getfiletype('Test.xlsx') == 'excel')

    def test_tblutil_getFileType_CSV(self):
        self.assertTrue(getfiletype('Test.csv') == 'csv')

    def test_tblutil_getFileType_other(self):
        with self.assertRaises(InvalidFileType):
            getfiletype('Test.txt')

    def test_tblutil_cvtColsStrToSet_err(self):
        with self.assertRaises(InvalidExcelColumn):
            cvtcolsstrtoset('A1')

    def test_tblutil_cvtColsStrToSet_num(self):
        result = cvtcolsstrtoset('1,3,7')
        self.assertTrue(result-{1, 2, 3})

    def test_tblutil_cvtColsStrToSet_alpha(self):
        result = cvtcolsstrtoset('A,C,G')
        self.assertTrue(result-{1, 2, 3})

    def test_tblutil_extractxltable_no_columns(self):
        try:
            wb = openpyxl.load_workbook('/users/rrehders/test/test.xlsx', data_only=True)
        except UserWarning as warn:
            pass
        xlsheet = wb.worksheets[0]
        result = extractxltable(xlsheet)
        self.assertTrue(len(result[0]) == xlsheet.max_column)

    def test_tblutil_extractxltable_three_columns(self):
        try:
            wb = openpyxl.load_workbook('/users/rrehders/test/test.xlsx', data_only=True)
        except UserWarning as warn:
            pass
        xlsheet = wb.worksheets[0]
        result = extractxltable(xlsheet, {1, 2, 3})
        self.assertTrue(len(result[0]) == 3)

    def test_tblutil_extractlisttable_no_columns(self):
        table = [[1, 2, 3], [4, 5, 6], [7, 8, 9]]
        result = extractlisttable(table)
        self.assertTrue(len(result[0]) == len(table[0]))

    def test_tblutil_extractlisttable_three_columns(self):
        table = [[1, 2, 3], [4, 5, 6], [7, 8, 9]]
        result = extractlisttable(table, {1, 2, 3})
        self.assertTrue(len(result[0]) == 3)

    def test_tblutil_tocsv_fail(self):
        self.assertFalse(tocsv('/users/rrehders/test/test.xls', 0))

    def test_tblutil_tocsv_success_minimum_parameter(self):
        self.assertTrue(tocsv('/users/rrehders/test/test.xlsx', 0))
        os.system('rm ./*.csv')

    def test_tblutil_tocsv_success_parameters(self):
        self.assertTrue(tocsv('/users/rrehders/test/test.xlsx', 0, cvtcolsstrtoset('A,B')))

    def test_tblutils_extractcols_fail_bad_file(self):
        with self.assertRaises(InvalidFileType):
            extractcols('/users/rrehders/test/fail.txt', cvtcolsstrtoset('A'))

    def test_tblutils_extractcols_fail_no_cols(self):
        with self.assertRaises(TypeError):
            extractcols('/users/rrehders/test/test.csv')

    def test_tblutils_extractcols_fail_empty_cols(self):
        self.assertFalse(extractcols('/users/rrehders/test/test.csv', set()))

    def test_tblutils_extractcols_success_csv(self):
        self.assertTrue(extractcols('/users/rrehders/test/test.csv', cvtcolsstrtoset('A,B')))
        os.system('rm ./*.csv')

    def test_tblutils_extractcols_success_xlsx(self):
        self.assertTrue(extractcols('/users/rrehders/test/test.xlsx', cvtcolsstrtoset('A,B'), 0))
        os.system('rm ./*.xlsx')

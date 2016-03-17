from tblutil import InvalidFileType, InvalidExcelColumn
from tblutil import getfiletype, cvtcolsstr, extractxltable, extractlisttable
from tblutil import cvtjoinstrindex, outcsvtable
import unittest
import openpyxl
import csv
import os


class SubFunctionTestCases(unittest.TestCase):
    def test_tblutil_getFileType_XLSX(self):
        self.assertTrue(getfiletype('Test.xlsx') == 'excel')

    def test_tblutil_getFileType_CSV(self):
        self.assertTrue(getfiletype('Test.csv') == 'csv')

    def test_tblutil_getFileType_other(self):
        with self.assertRaises(InvalidFileType):
            getfiletype('Test.txt')

    def test_tblutil_cvtcolsstrtoset_err(self):
        with self.assertRaises(InvalidExcelColumn):
            cvtcolsstr('A1')

    def test_tblutil_cvtcolsstrtoset_single_num(self):
        result = cvtcolsstr('1')
        self.assertTrue(result==1)

    def test_tblutil_cvtcolsstrtoset_single_alpha(self):
        result = cvtcolsstr('A')
        self.assertTrue(result==1)

    def test_tblutil_cvtcolsstrtoset_multiple_num(self):
        result = cvtcolsstr('1,3,7')
        self.assertFalse(result-{1, 3, 7})

    def test_tblutil_cvtcolsstrtoset_multiple_alpha(self):
        result = cvtcolsstr('A,C,G')
        self.assertFalse(result-{1, 3, 7})
        
    def test_tblutil_cvtstrindextoset_single_num(self):
        result = cvtjoinstrindex('1')
        self.assertTrue(result==(1,1))

    def test_tblutil_cvtstrindextoset_single_alpha(self):
        result = cvtjoinstrindex('A')
        self.assertTrue(result==(1,1))

    def test_tblutil_cvtstrindextoset_double_num(self):
        result = cvtjoinstrindex('2,3')
        self.assertFalse(result==(2, 3))

    def test_tblutil_cvtstrindextoset_double_alpha(self):
        result = cvtjoinstrindex('C,B')
        self.assertFalse(result==(2, 3))

    def test_tblutil_extractxltable_no_columns(self):
        try:
            wb = openpyxl.load_workbook('/users/rrehders/test/test.xlsx', data_only=True)
        except UserWarning as warn:
            pass
        xlsheet = wb.worksheets[0]
        result = extractxltable(xlsheet)
        self.assertTrue(len(result[0]) == xlsheet.max_column)

    def test_tblutil_extractxltable_three_columns(self):
        try:
            wb = openpyxl.load_workbook('/users/rrehders/test/test.xlsx', data_only=True)
        except UserWarning as warn:
            pass
        xlsheet = wb.worksheets[0]
        result = extractxltable(xlsheet, {1, 2, 3})
        self.assertTrue(len(result[0]) == 3)

    def test_tblutil_extractlisttable_no_columns(self):
        table = [[1, 2, 3], [4, 5, 6], [7, 8, 9]]
        result = extractlisttable(table)
        self.assertTrue(len(result[0]) == len(table[0]))

    def test_tblutil_extractlisttable_three_columns(self):
        table = [[1, 2, 3], [4, 5, 6], [7, 8, 9]]
        result = extractlisttable(table, {1, 2, 3})
        self.assertTrue(len(result[0]) == 3)

    def test_tblutil_outcsvtable(self):
        table = [[1, 2, 3], [4, 5, 6], [7, 8, 9]]
        result = outcsvtable(table, '/users/rrehders/test/result.csv', 'subitem')
        with open('result-subitem.csv', 'r') as filein:
            csvin = csv.reader(filein)
            data = [[int(item) for item in row] for row in csvin]
        self.assertTrue(table == data)
        os.system('rm ./*.csv')
